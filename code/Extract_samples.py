from openpyxl import Workbook

def Extract(output,time,out,jump,arrival):
    time.append(output[0])
    out+=[data for data in output[3].values()]

    for i in range(len(output[4])):
        if(len(jump)==i):
            jump.append([])
            jump[i].append(output[4][i])
            continue
        jump[i].append(output[4][i])

    arrival_=[]
    for i in range(len(output[2])):
        value=[data for data in output[2][i].values()]
        arrival_.append(value)

    for i in range(len(output[4])):
        if(len(arrival)==i):
            arrival.append(arrival_[i])
            continue
        arrival[i]+=arrival_[i]

    return (time,out,jump,arrival)




def Copy_Xslx(time,out,jump,arrival,path):

    work=Workbook()
    calc=work.active

    for i, valor in enumerate(time,start=1):
        calc.cell(row=i+1, column=1).value = valor

    for i, valor in enumerate(out,start=1):
        calc.cell(row=i+1, column=2).value = valor

    for i, lista in enumerate(jump, start=1):
        for j, valor in enumerate(lista, start=1):
            calc.cell(row=j+1, column=i + 2).value = valor

    for i, lista in enumerate(arrival, start=1):
        for j, valor in enumerate(lista, start=1):
            calc.cell(row=j+1, column=i + 2+len(jump)).value = valor


    work.save(path)